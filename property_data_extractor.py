"""
Property data extraction module for PropIntel.

This module provides utilities for extracting property data from various sources and formats,
such as Excel files, and loading them into the database.
"""

def get_db_config():
    """
    Get database configuration from environment variable (for Heroku)
    or from config file (for local development)
    """
    import os
    import configparser
    from urllib.parse import urlparse
    
    # Check for DATABASE_URL environment variable (set by Heroku)
    database_url = os.environ.get('DATABASE_URL')
    
    if database_url:
        # Parse Heroku DATABASE_URL
        # Note: Heroku uses 'postgres://' but psycopg2 needs 'postgresql://'
        if database_url.startswith('postgres://'):
            database_url = database_url.replace('postgres://', 'postgresql://', 1)
            
        # Parse the URL
        result = urlparse(database_url)
        
        # Build connection parameters
        return {
            "user": result.username,
            "password": result.password,
            "host": result.hostname,
            "port": result.port or 5432,
            "database": result.path[1:],  # Remove leading slash
        }
    else:
        # Fallback to config file for local development
        config = configparser.ConfigParser()
        
        # Default connection parameters
        default_params = {
            "user": "u15p78tmoefhv2",
            "password": "p78dc6c2370076ee1ac7f23f370d707687e8400f94032cccdb35ddd1d7b37381f",
            "host": "c1i13pt05ja4ag.cluster-czrs8kj4isg7.us-east-1.rds.amazonaws.com",
            "port": 5432,
            "database": "d1oncga6g47frr",
        }

        # Try to read from config file
        if os.path.exists('db_config.ini'):
            try:
                config.read('db_config.ini')
                if 'database' in config:
                    return {
                        "user": config['database'].get('user', default_params['user']),
                        "password": config['database'].get('password', default_params['password']),
                        "host": config['database'].get('host', default_params['host']),
                        "port": int(config['database'].get('port', default_params['port'])),
                        "database": config['database'].get('database', default_params['database']),
                    }
            except Exception as e:
                print(f"Error reading config file: {e}. Using default parameters.")
        
        return default_params

import pandas as pd
import psycopg2
from geopy.geocoders import Nominatim
import openpyxl
from datetime import datetime, timedelta
import re
import time
import logging
import os
import sys

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    filename='propintel_import.log',
    filemode='a'
)
logger = logging.getLogger('propintel.data_extractor')

# Add console handler
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

def get_db_connection():
    """Get a database connection using the provided configuration"""
    import os
    from urllib.parse import urlparse
    
    # Check if running on Heroku
    if 'DATABASE_URL' in os.environ:
        # Parse database URL for Heroku
        database_url = os.environ['DATABASE_URL']
        # Check for postgres:// prefix and replace with postgresql://
        if database_url.startswith('postgres://'):
            database_url = database_url.replace('postgres://', 'postgresql://', 1)
        conn = psycopg2.connect(database_url, sslmode='require')
    else:
        # Try to load from .env file first
        try:
            from dotenv import load_dotenv
            load_dotenv()
            
            db_host = os.environ.get('DB_HOST')
            db_name = os.environ.get('DB_NAME')
            db_user = os.environ.get('DB_USER')
            db_password = os.environ.get('DB_PASSWORD')
            db_port = os.environ.get('DB_PORT', '5432')
            
            if db_host and db_name and db_user and db_password:
                conn = psycopg2.connect(
                    user=db_user,
                    password=db_password,
                    host=db_host,
                    port=db_port,
                    dbname=db_name
                )
                logger.info("Connected using .env configuration")
                conn.autocommit = False
                return conn
        except Exception as e:
            logger.warning(f"Failed to load from .env: {e}")
        
        # Try to load from config file
        try:
            import configparser
            config = configparser.ConfigParser()
            config.read('db_config.ini')
            
            conn = psycopg2.connect(
                user=config['database']['user'],
                password=config['database']['password'],
                host=config['database']['host'],
                port=config['database']['port'],
                dbname=config['database']['database']
            )
            logger.info("Connected using db_config.ini")
        except Exception as e:
            logger.warning(f"Failed to load config: {e}")
            # Use the database parameters from db_connect.py default values
            conn = psycopg2.connect(
                user="u15p78tmoefhv2",
                password="p78dc6c2370076ee1ac7f23f370d707687e8400f94032cccdb35ddd1d7b37381f",
                host="c1i13pt05ja4ag.cluster-czrs8kj4isg7.us-east-1.rds.amazonaws.com",
                port="5432",
                database="d1oncga6g47frr"
            )
            logger.info("Connected using fallback credentials")
    
    conn.autocommit = False
    return conn

def is_valid_date(date_value):
    """Check if a value is a valid date and not a header or summary row"""
    if date_value is None:
        return False
        
    # Check if it's already a date object
    if isinstance(date_value, datetime) or isinstance(date_value, datetime.date):
        return True
        
    # Skip common summary row labels
    if isinstance(date_value, str):
        skip_terms = ['total', 'totals', 'profit', 'margin', 'sum', 'average', 'header']
        if any(term in date_value.lower() for term in skip_terms):
            return False
        
        # Check if it's a header label (all caps, no numbers)
        if date_value.upper() == date_value and not any(c.isdigit() for c in date_value) and len(date_value) > 3:
            return False
            
    # Handle Excel numeric dates
    if isinstance(date_value, (int, float)) and 5000 < date_value < 50000:  # Reasonable date range
        try:
            # Test if it can be converted to a date
            datetime(1899, 12, 30) + timedelta(days=int(date_value))
            return True
        except:
            pass
            
    return False

def format_date(date_value):
    """Format a date value to a standard format, with validation"""
    if not is_valid_date(date_value):
        return None
        
    try:
        if isinstance(date_value, str):
            # Try different date formats
            if re.match(r'^\d{1,2}/\d{1,2}/\d{2,4}$', date_value):
                return datetime.strptime(date_value, '%d/%m/%y').date()
            elif re.match(r'^\d{2,4}-\d{1,2}-\d{1,2}$', date_value):
                return datetime.strptime(date_value, '%Y-%m-%d').date()
            else:
                logger.warning(f"Unrecognized date format: {date_value}")
                return None
        elif isinstance(date_value, (int, float)):
            # Excel dates are stored as days since 1900-01-01
            return (datetime(1899, 12, 30) + timedelta(days=int(date_value))).date()
        elif isinstance(date_value, (datetime, datetime.date)):
            # Already a date or datetime object
            if isinstance(date_value, datetime):
                return date_value.date()
            return date_value
    except Exception as e:
        logger.warning(f"Could not convert to date: {date_value}, Error: {e}")
        
    return None

def is_currency(value):
    """Check if a value represents currency"""
    if value is None:
        return False
    
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return True
        
    if isinstance(value, str):
        # Remove currency symbols, commas and spaces
        clean_value = value.replace('$', '').replace(',', '').replace(' ', '')
        
        # Try to convert to float
        try:
            float(clean_value)
            return True
        except ValueError:
            return False
            
    return False

def format_currency(value):
    """Format a value as currency"""
    if not is_currency(value):
        return None
        
    if isinstance(value, (int, float)):
        return float(value)
        
    if isinstance(value, str):
        # Remove currency symbols, commas and spaces
        clean_value = value.replace('$', '').replace(',', '').replace(' ', '')
        
        # Try to convert to float
        try:
            return float(clean_value)
        except ValueError:
            return None

def geocode_address(address, retries=3, delay=1):
    """Convert address to coordinates using OpenStreetMap Nominatim"""
    if not address or not isinstance(address, str):
        return None, None
        
    # Clean the address
    address = address.strip()
    
    for attempt in range(retries):
        try:
            geolocator = Nominatim(user_agent="propintel_geocoder")
            location = geolocator.geocode(address)
            
            if location:
                return location.latitude, location.longitude
                
            # Add ", Australia" if not found and not already there
            if ", Australia" not in address:
                location = geolocator.geocode(address + ", Australia")
                if location:
                    return location.latitude, location.longitude
                    
            # Last attempt with "VIC, Australia"
            if ", VIC" not in address and ", Victoria" not in address:
                location = geolocator.geocode(address + ", VIC, Australia")
                if location:
                    return location.latitude, location.longitude
                    
            return None, None
            
        except Exception as e:
            logger.warning(f"Geocoding error for {address}: {e}")
            if attempt < retries - 1:
                logger.info(f"Retrying in {delay} seconds...")
                time.sleep(delay)
                
    return None, None

def extract_property_data_from_excel(file_path):
    """Extract property data from an Excel file"""
    if not os.path.exists(file_path):
        logger.error(f"File not found: {file_path}")
        return None
        
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # Find data in the first sheet
        sheet = workbook.active
        
        # Initialize data containers
        properties = []
        
        # Find headers and data
        headers = {}
        data_row_start = 0
        
        # Scan for headers in first 20 rows
        for row_idx in range(1, min(20, sheet.max_row + 1)):
            potential_headers = {}
            
            for col_idx in range(1, min(30, sheet.max_column + 1)):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                
                if not cell_value:
                    continue
                    
                if isinstance(cell_value, str):
                    clean_header = cell_value.strip().lower()
                    
                    # Common property field names
                    if any(term in clean_header for term in ['property', 'address', 'location']):
                        potential_headers[clean_header] = col_idx
                        
            if len(potential_headers) >= 2:  # Found at least two potential property-related headers
                headers = potential_headers
                data_row_start = row_idx + 1
                break
                
        if not headers or data_row_start == 0:
            logger.error(f"Could not find property headers in {file_path}")
            return None
            
        # Map common header variations
        header_mapping = {
            'property': ['property', 'property name', 'name', 'title'],
            'address': ['address', 'property address', 'location', 'street'],
            'purchase_date': ['date', 'purchase date', 'date purchased', 'purchase'],
            'project_manager': ['manager', 'project manager', 'pm', 'property manager'],
            'notes': ['notes', 'description', 'comments', 'details']
        }
        
        # Find header columns based on mapping
        column_mapping = {}
        for field, variations in header_mapping.items():
            for variation in variations:
                for header, col_idx in headers.items():
                    if variation in header:
                        column_mapping[field] = col_idx
                        break
        
        # Find date columns and amount columns for auto-detection
        date_candidates = set()
        amount_candidates = set()
        
        # Scan the first few data rows to detect date and amount columns
        for row_idx in range(data_row_start, min(data_row_start + 5, sheet.max_row + 1)):
            for col_idx in range(1, min(30, sheet.max_column + 1)):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if not cell_value:
                    continue
                
                # Check for dates
                if isinstance(cell_value, datetime):
                    date_candidates.add(col_idx)
                elif isinstance(cell_value, str) and re.match(r'\d{1,2}/\d{1,2}/\d{2,4}', cell_value):
                    date_candidates.add(col_idx)
                    
                # Check for amounts
                if isinstance(cell_value, (int, float)) and not isinstance(cell_value, bool):
                    amount_candidates.add(col_idx)
                elif isinstance(cell_value, str):
                    try:
                        amount = format_currency(cell_value)
                        if amount is not None:
                            amount_candidates.add(col_idx)
                    except:
                        pass
        
        # Add detected date columns to mapping if not already set
        if 'purchase_date' not in column_mapping and date_candidates:
            column_mapping['purchase_date'] = min(date_candidates)  # Just pick the first date column
            
        # Extract property data
        for row_idx in range(data_row_start, sheet.max_row + 1):
            # Check if this is a data row
            if all(sheet.cell(row=row_idx, column=col_idx).value is None for col_idx in column_mapping.values()):
                continue
                
            # Extract property fields
            property_data = {}
            
            for field, col_idx in column_mapping.items():
                value = sheet.cell(row=row_idx, column=col_idx).value
                
                # Special handling for dates
                if field == 'purchase_date':
                    value = format_date(value)
                
                property_data[field] = value if value is not None else ""
                
            # Skip if no property name or address
            if not property_data.get('property') and not property_data.get('address'):
                continue
                
            # Add to properties list
            properties.append(property_data)
            
        logger.info(f"Extracted {len(properties)} properties from {file_path}")
        return properties
        
    except Exception as e:
        logger.error(f"Error extracting property data from {file_path}: {e}")
        return None

def import_properties_to_db(properties, user_id):
    """Import property data to the database"""
    if not properties:
        logger.error("No properties to import")
        return False
        
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            # First, check if properties table exists
            cur.execute("""
                SELECT EXISTS (
                    SELECT FROM pg_tables
                    WHERE schemaname = 'propintel'
                    AND tablename  = 'properties'
                );
            """)
            table_exists = cur.fetchone()[0]
            
            if not table_exists:
                logger.error("Properties table does not exist in the database")
                return False
            
            # Import each property
            property_ids = []
            
            for prop in properties:
                # Get property data
                property_name = prop.get('property', '').strip()
                address = prop.get('address', '').strip()
                project_manager = prop.get('project_manager', '').strip()
                purchase_date = prop.get('purchase_date')
                notes = prop.get('notes', '').strip()
                
                # Skip if no name or address
                if not property_name and not address:
                    continue
                    
                # Use address as property name if missing
                if not property_name:
                    property_name = address
                
                # Geocode address to get coordinates
                latitude, longitude = geocode_address(address)
                
                # Insert property
                cur.execute("""
                    INSERT INTO propintel.properties 
                    (user_id, property_name, address, project_manager, purchase_date, notes, latitude, longitude)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING property_id
                """, (
                    user_id,
                    property_name,
                    address,
                    project_manager,
                    purchase_date,
                    notes,
                    latitude,
                    longitude
                ))
                
                property_id = cur.fetchone()[0]
                property_ids.append(property_id)
                
                logger.info(f"Imported property {property_name} with ID {property_id}")
                
            conn.commit()
            logger.info(f"Successfully imported {len(property_ids)} properties")
            return property_ids
            
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Error importing properties to database: {e}")
        return False
    finally:
        if conn:
            conn.close()

def extract_finances_from_excel(file_path, property_map=None):
    """
    Extract financial data (income & expenses) from Excel file.
    
    Args:
        file_path: Path to the Excel file
        property_map: Optional dictionary mapping property names to property IDs
        
    Returns:
        Dictionary with income and expense records
    """
    if not os.path.exists(file_path):
        logger.error(f"File not found: {file_path}")
        return None
        
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # Find finances in the workbook
        income_records = []
        expense_records = []
        
        # Try to find income and expense sheets
        sheet_names = workbook.sheetnames
        income_keywords = ['income', 'revenue', 'inbound', 'inflow', 'incoming']
        expense_keywords = ['expense', 'cost', 'payment', 'outbound', 'outflow', 'outgoing']
        
        income_sheet_candidates = []
        expense_sheet_candidates = []
        
        # Find sheets that might contain income or expense data
        for sheet_name in sheet_names:
            lower_name = sheet_name.lower()
            
            if any(kw in lower_name for kw in income_keywords):
                income_sheet_candidates.append(sheet_name)
                
            if any(kw in lower_name for kw in expense_keywords):
                expense_sheet_candidates.append(sheet_name)
                
        # If no specific sheets found, try the first sheet
        if not income_sheet_candidates and not expense_sheet_candidates:
            process_generic_finance_sheet(workbook.active, income_records, expense_records, property_map)
        else:
            # Process income sheets
            for sheet_name in income_sheet_candidates:
                sheet = workbook[sheet_name]
                process_income_sheet(sheet, income_records, property_map)
                
            # Process expense sheets
            for sheet_name in expense_sheet_candidates:
                sheet = workbook[sheet_name]
                process_expense_sheet(sheet, expense_records, property_map)
                
        logger.info(f"Extracted {len(income_records)} income records and {len(expense_records)} expense records from {file_path}")
        return {
            'income': income_records,
            'expenses': expense_records
        }
        
    except Exception as e:
        logger.error(f"Error extracting finances from {file_path}: {e}")
        return None

def process_generic_finance_sheet(sheet, income_records, expense_records, property_map=None):
    """Process a generic sheet that might contain both income and expenses"""
    # Try to identify headers
    headers = {}
    data_row_start = 0
    
    # Define common header keywords
    property_keywords = ['property', 'address', 'location', 'site']
    date_keywords = ['date', 'day', 'time']
    amount_keywords = ['amount', 'value', 'total', 'sum', 'payment', 'price']
    description_keywords = ['description', 'details', 'notes', 'comments', 'info']
    category_keywords = ['category', 'type', 'class', 'group']
    
    # Scan for headers
    for row_idx in range(1, min(20, sheet.max_row + 1)):
        row_headers = {}
        
        for col_idx in range(1, min(20, sheet.max_column + 1)):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            
            if not cell_value or not isinstance(cell_value, str):
                continue
                
            header_text = cell_value.strip().lower()
            
            # Check for property identifier
            if any(kw in header_text for kw in property_keywords):
                row_headers['property'] = col_idx
                
            # Check for date
            elif any(kw in header_text for kw in date_keywords):
                row_headers['date'] = col_idx
                
            # Check for amount
            elif any(kw in header_text for kw in amount_keywords):
                row_headers['amount'] = col_idx
                
            # Check for description
            elif any(kw in header_text for kw in description_keywords):
                row_headers['description'] = col_idx
                
            # Check for category
            elif any(kw in header_text for kw in category_keywords):
                row_headers['category'] = col_idx
                
        # If we found at least date and amount, we can work with this
        if 'date' in row_headers and 'amount' in row_headers:
            headers = row_headers
            data_row_start = row_idx + 1
            break
            
    if not headers or data_row_start == 0:
        logger.warning(f"Could not find finance headers in sheet {sheet.title}")
        return
        
    # Process data rows
    income_count = 0
    expense_count = 0
    
    for row_idx in range(data_row_start, sheet.max_row + 1):
        # Get date and amount
        date_value = sheet.cell(row=row_idx, column=headers['date']).value
        amount_value = sheet.cell(row=row_idx, column=headers['amount']).value
        
        # Skip empty rows
        if date_value is None or amount_value is None:
            continue
            
        # Format date
        date = format_date(date_value)
        if date is None:
            continue
            
        # Format amount
        amount = format_currency(amount_value)
        if amount is None:
            continue
            
        # Get property identifier
        property_name = None
        if 'property' in headers:
            property_value = sheet.cell(row=row_idx, column=headers['property']).value
            if property_value:
                property_name = str(property_value).strip()
                
        # Get property_id from map
        property_id = None
        if property_map and property_name:
            for name, pid in property_map.items():
                if name.lower() in property_name.lower() or property_name.lower() in name.lower():
                    property_id = pid
                    break
                    
        # Get description
        description = ""
        if 'description' in headers:
            description_value = sheet.cell(row=row_idx, column=headers['description']).value
            if description_value:
                description = str(description_value).strip()
                
        # Get category
        category = None
        if 'category' in headers:
            category_value = sheet.cell(row=row_idx, column=headers['category']).value
            if category_value:
                category = str(category_value).strip().lower()
                
        # Determine if income or expense
        is_income = amount > 0 or (category and any(term in category for term in ['income', 'revenue', 'inflow']))
        
        # Create record
        record = {
            'date': date,
            'amount': abs(amount),
            'property': property_name,
            'property_id': property_id,
            'description': description,
            'category': category
        }
        
        # Add to appropriate list
        if is_income:
            income_records.append(record)
            income_count += 1
        else:
            expense_records.append(record)
            expense_count += 1
            
    logger.info(f"Processed generic finance sheet: {income_count} income records, {expense_count} expense records")

def process_income_sheet(sheet, income_records, property_map=None):
    """Process a sheet containing income data"""
    # Similar logic to process_generic_finance_sheet but focused on income
    # Implementation would extract income records and add them to income_records list
    pass

def process_expense_sheet(sheet, expense_records, property_map=None):
    """Process a sheet containing expense data"""
    # Similar logic to process_generic_finance_sheet but focused on expenses
    # Implementation would extract expense records and add them to expense_records list
    pass

def import_finances_to_db(financial_data, user_id):
    """Import financial data to the database"""
    if not financial_data:
        logger.error("No financial data to import")
        return False
        
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            # Import income records
            income_count = 0
            for record in financial_data.get('income', []):
                # Get data
                income_date = record.get('date')
                income_amount = record.get('amount')
                property_id = record.get('property_id')
                income_source = record.get('category', '').strip()
                income_details = record.get('description', '').strip()
                
                # Skip if missing required fields
                if not income_date or not income_amount:
                    continue
                    
                # Insert income record
                cur.execute("""
                    INSERT INTO propintel.money_in
                    (user_id, property_id, income_source, income_details, income_date, income_amount)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (
                    user_id,
                    property_id,
                    income_source or "other",
                    income_details,
                    income_date,
                    income_amount
                ))
                
                income_count += 1
                
            # Import expense records
            expense_count = 0
            for record in financial_data.get('expenses', []):
                # Get data
                expense_date = record.get('date')
                expense_amount = record.get('amount')
                property_id = record.get('property_id')
                expense_category = record.get('category', '').strip().lower()
                expense_details = record.get('description', '').strip()
                
                # Skip if missing required fields
                if not expense_date or not expense_amount:
                    continue
                    
                # Auto-categorize if needed
                if not expense_category:
                    lower_details = expense_details.lower()
                    if 'wage' in lower_details or 'salary' in lower_details:
                        expense_category = 'wage'
                    elif 'manager' in lower_details or 'pm ' in lower_details:
                        expense_category = 'project_manager'
                    elif 'material' in lower_details or 'supplies' in lower_details:
                        expense_category = 'material'
                    else:
                        expense_category = 'miscellaneous'
                        
                # Insert expense record
                cur.execute("""
                    INSERT INTO propintel.money_out
                    (user_id, property_id, expense_category, expense_details, expense_date, expense_amount)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (
                    user_id,
                    property_id,
                    expense_category or "miscellaneous",
                    expense_details,
                    expense_date,
                    expense_amount
                ))
                
                expense_count += 1
                
            conn.commit()
            logger.info(f"Successfully imported {income_count} income records and {expense_count} expense records")
            return {
                'income_count': income_count,
                'expense_count': expense_count
            }
            
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Error importing finances to database: {e}")
        return False
    finally:
        if conn:
            conn.close()

def extract_work_data_from_excel(file_path, property_map=None):
    """Extract work record data from Excel"""
    # Implementation similar to extract_finances_from_excel but for work records
    pass

def import_work_to_db(work_data, user_id):
    """Import work records to the database"""
    # Implementation similar to import_finances_to_db but for work records
    pass

# If run directly, perform a test extraction
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python property_data_extractor.py <excel_file_path>")
        sys.exit(1)
        
    file_path = sys.argv[1]
    print(f"Testing extraction from {file_path}")
    
    # Extract property data
    properties = extract_property_data_from_excel(file_path)
    
    if properties:
        print(f"Extracted {len(properties)} properties")
        for i, prop in enumerate(properties[:5]):
            print(f"Property {i+1}: {prop.get('property', 'Unknown')} - {prop.get('address', 'No address')}")
    else:
        print("No property data found")